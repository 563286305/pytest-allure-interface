# -*- coding: utf-8 -*-
'''
Created on 2018.02.02
@author: wuyou
'''

from openpyxl import Workbook, load_workbook
from contextlib import closing
import re
import sys
from openpyxl.styles import Font, Color, Alignment
from openpyxl.styles import colors



class operExcel():

	#读取两个指定列，并组成字典
	def get_excel_name_value_dict(self, file_name, sheet_name,colname1,colname2,row_start,row_end):
		namelist = self.colname_read_excel_return_list(file_name,sheet_name,colname1,row_start, row_end)
		valuelist = self.colname_read_excel_return_list(file_name,sheet_name,colname2,row_start, row_end)
		return dict(list(zip(namelist, valuelist)))

	# 按列分段读取Excel
	def colname_read_excel_return_list(self, file_name, sheet_name, col_name, start, end):
		with closing(load_workbook(filename=file_name, data_only=True)) as wb:
			ws = wb[str(sheet_name)]

			if int(start) == int(end):
				cellname = str(col_name) + str(start)
				cellValue = ws[str(cellname)].value
				cellValueList = [str(cellValue)]
			else:
				cellValueList = []
				for i in range(int(start), int(end)+1):
					cellname = str(col_name) + str(i)
					cellValue = ws[str(cellname)].value
					cellValueList.append(str(cellValue))
			return cellValueList


	# 获取指定sheet中行数
	def get_excel_row_count(self, file_name, sheet_name):
		with closing(load_workbook(filename=file_name)) as wb:
#			rows = wb.get_sheet_by_name(name=str(sheet_name)).max_row
			rows = wb[str(sheet_name)].max_row

			return rows

	#获取某列(为标记列所使用)的index，并name：index存在字典中
	def get_cell_index(self,file_name, sheet_name, rowNo):
		with closing(load_workbook(filename=file_name)) as wb:
			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			num = int(rowNo) - 1
			tagdict = {}
			for cell in list(ws.rows)[num]:
				if cell.value != None:
					value = cell.value
					value = str(value)
#					print "value:    "+ value
					index = cell.coordinate
#					print "index: " + index
					indexlist = re.findall('[A-Z]', index)
					colindex = ''.join(indexlist)
				tagdict[value] = colindex
		return 	tagdict

	def add_result_bytag(self,file_name, sheet_name, tagcolumnNo, dictlist):
		#excel表中得到tag列的 index，存在字典
		tagdict = self.get_cell_index(file_name, sheet_name, tagcolumnNo)
		# print 'tagdict: '
		# print tagdict
		index_result_dict = {}

		#写入数据字典list长度范围内
		for i in range(len(dictlist)):

			#去每个字典里取得tag名,result结果
			tagname = dictlist[i]['tag']
			result = dictlist[i]['result']
			# print 'tagname :' + tagname
			# print 'result :' + result
			#根据tag名去excel得到的字典中，取得对应的index
			tagindex = tagdict[tagname]
			# print 'tagindex: '+ tagindex
			#拆分index（E2只取2,G23只取23）
			tagindexlist = re.findall('[0-9]', tagindex)
			indextmp = ''.join(tagindexlist)
			result_index = 'H'+ indextmp
			# print 'resultindex: '+ result_index

			index_result_dict[result_index]=result

		# print index_result_dict
		#
		with closing(load_workbook(filename=file_name)) as wb:
			ws = wb[sheet_name]
			ft1 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft2 = Font(name="微软雅黑", color=colors.GREEN,size=10, bold=True)
			ft3 = Font(name="微软雅黑 Light", size=9)
			for key in list(index_result_dict.keys()):
				ws[key] = index_result_dict[key]
				# print "**********"
				# print ws['H2']
				if index_result_dict[key] == 'FAIL':
#					ws.cell(key).font = ft1
					ws[key].font = ft1
				if index_result_dict[key] == 'PASS':
#					ws.cell(key).font = ft2
					ws[key].font = ft2

			wb.save(file_name)

	def copy_excel(self,openfile,savefile):
		with closing(load_workbook(filename=openfile)) as wb:
			wb.save(savefile)

	# 将二维list内容按列写入到Excel
	def dictList_write_excel_by_dict_key_batch(self, file_name, sheet_name, row_start, col_list, value_list_list):
		with closing(load_workbook(filename=file_name)) as wb:

			sheets = wb.sheetnames
			if sheet_name in sheets:
				ws = wb[str(sheet_name)]
			else:
				ws = wb.create_sheet(str(sheet_name))
			#ft1 = Font(name="微软雅黑 Light", size=10, bold=False)
			ft1 = Font(name="微软雅黑", size=10, bold=True)
			ft2 = Font(name="微软雅黑", color=colors.RED, size=10)
			ft3 = Font(name="微软雅黑", size=10)
			for i in range(len(value_list_list)):
				for j in range(len(value_list_list[i])):
					cellname = col_list[j] + str(int(row_start)+i)
					ws[str(cellname)] = value_list_list[i][j]
					print(value_list_list[i][j])
					if str(value_list_list[i][j])=="PASS":
						ws[str(cellname)].font = ft1
					if str(value_list_list[i][j])=="FAIL":
						ws[str(cellname)].font = ft2
					else:
						ws[str(cellname)].font = ft3
		wb.save(file_name)


#Just for test
if __name__ == '__main__':
	handle = ReadWriteExcel()
	templatefile ='..//Template//Case_Flow_Base_Check.xlsx'
	filename = '..\Report\Excel\Case_Flow_Base_Check.xlsx'
	sheet_name = 'flow_base'
	handle.get_cell_index(filename,sheet_name,'7')
	dictlist=[{'tag':'flow1','result':'PASS'},{'tag':'flow2','result':'FAIL'}]
	handle.add_result_bytag( filename, sheet_name,'7', dictlist)
#	handle.copy_excel(templatefile,filename)